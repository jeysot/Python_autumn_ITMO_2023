import openpyxl
wb = openpyxl.Workbook()
ws1 = wb.active

ws1['A1'].value = 'Фамилия'
ws1['B1'].value = 'Результат'

ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Сидоров'
ws1['A5'].value = 'Жуков'
ws1['A6'].value = 'Андреев'

ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'
ws1['B5'].value = '200'
ws1['B5'].value = '75'
ws1['B6'].value = '115'

wb.create_sheet('Лист2')
ws2 = wb['Лист2']

ws2['A1'].value = 'Минимальное'
ws2['A2'].value = 'Максимальное'
ws2['A3'].value = 'Сред. арефм.'
ws2['A4'].value = 'Медиана'

result = []
for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        result.append(int(work))

ws2['B1'].value = min(result)
ws2['B2'].value = max(result)
ws2['B3'].value = sum(result)/len(result)
result.sort()

if len(result) % 2 != 0:
    med = result[len(result)//2]
    ws2['B4'].value = med
else:
    med = (result[len(result) // 2 - 1] + result[len(result) // 2]) / 2
    ws2['B4'].value = med


wb.save('Task10_3.xlsx')
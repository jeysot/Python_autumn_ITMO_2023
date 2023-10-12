import openpyxl

wb = openpyxl.Workbook()


ws1 = wb.create_sheet('Лист1')
ws1['A1'].value = 'Фамилия'
ws1['B1'].value = 'Результат'
ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Сидоров'
ws1['A5'].value = 'Жуков'
ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'
ws1['B5'].value = '200'


wb.create_sheet('Лист2')
ws2 = wb['Лист2']
ws2['A1'].value = 'Фамилия'
ws2['B1'].value = 'Результат'
ws2['A2'].value = 'Андреев'
ws2['A3'].value = 'Петров'
ws2['A4'].value = 'Сидоров'
ws2['A5'].value = 'Жуков'
ws2['B2'].value = '150'
ws2['B3'].value = '400'
ws2['B4'].value = '200'
ws2['B5'].value = '200'


wb.create_sheet('Лист3')
ws3 = wb['Лист3']
ws3['A1'].value = 'Фамилия'
ws3['B1'].value = 'Результат'

result = {}

for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)

for row in ws2.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'Фамилия':
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)

sorted_result = dict(sorted(result.items()))

num_row = 2
for name, work in sorted_result.items():
    ws3.cell(row=num_row, column=1, value=name)
    ws3.cell(row=num_row, column=2, value=work)
    num_row += 1

wb.save('Task10_2.xlsx')
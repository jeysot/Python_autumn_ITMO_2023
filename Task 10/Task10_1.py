import openpyxl

wb = openpyxl.load_workbook('Task10_1.xlsx')

wb.active
ws1 = wb['Лист1']

ws1['A1'].value = 'ФИО'
ws1['B1'].value = 'Результат'
ws1['A2'].value = 'Иванов'
ws1['A3'].value = 'Петров'
ws1['A4'].value = 'Иванов'
ws1['B2'].value = '100'
ws1['B3'].value = '400'
ws1['B4'].value = '200'

wb.create_sheet('Лист2')
ws2 = wb['Лист2']

result = {}

for row in ws1.iter_rows(values_only=True):
    name = row[0]
    work = row[1]
    if name != 'ФИО':
        if name in result:
            result[name] += int(work)
        else:
            result[name] = int(work)

ws2['A1'].value = 'ФИО'
ws2['B1'].value = 'Результат'

row_num = 2
for name, work in result.items():
    ws2.cell(row=row_num, column=1, value=name)
    ws2.cell(row=row_num, column=2, value=work)
    row_num += 1

ws2.cell(row=row_num, column=1, value='ИТОГО')
ws2.cell(row=row_num, column=2, value=sum(result.values()))

wb.save('Task10_1.xlsx')
